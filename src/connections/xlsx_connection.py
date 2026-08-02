"""
XlsxConnection — Service kết nối dữ liệu qua file XLSX local.

Thay thế hoàn toàn GoogleSheetsManager.
Giữ nguyên interface (method names & return types) để worker.py không cần sửa.

Cấu trúc file XLSX:
  Sheet "Outlooks" — danh sách email Outlook/Hotmail đầu vào (đọc + cập nhật status)
  Sheet "Gmails"   — danh sách email Gmail đầu vào (có otp_email, otp_pass)
  Sheet "Iclouds"  — danh sách email iCloud đầu vào (có otp_email, otp_pass)
  Sheet "Accounts" — kết quả đăng ký (ghi/upsert)
  Sheet "Proxies"  — danh sách proxy (tùy chọn)
"""

import threading
from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.utils.logger import get_logger
from src import config

log = get_logger("xlsx_connection")

def repair_xlsx_crc(corrupted_file_path: str, output_file_path: str):
    """
    Tự động sửa lỗi Bad CRC-32 trong file XLSX Zip archive.
    Tái tạo lại các file XML bên trong với checksum CRC32 hợp lệ.
    """
    import zipfile
    import zlib
    with zipfile.ZipFile(corrupted_file_path, 'r') as zin:
        with zipfile.ZipFile(output_file_path, 'w', compression=zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                content = None
                try:
                    fp = zin.open(item)
                    if hasattr(fp, '_expected_crc'):
                        fp._expected_crc = None
                    content = fp.read()
                except Exception:
                    try:
                        zinfo = zin.getinfo(item.filename)
                        with open(corrupted_file_path, 'rb') as f:
                            f.seek(zinfo.header_offset)
                            header = f.read(30)
                            fname_len = int.from_bytes(header[26:28], 'little')
                            extra_len = int.from_bytes(header[28:30], 'little')
                            f.seek(zinfo.header_offset + 30 + fname_len + extra_len)
                            compressed_data = f.read(zinfo.compress_size)
                            if zinfo.compress_type == zipfile.ZIP_DEFLATED:
                                content = zlib.decompress(compressed_data, -15)
                            else:
                                content = compressed_data
                    except Exception as ex:
                        log.warning(f"Không thể đọc file {item.filename} trong archive: {ex}")
                        continue

                if content is not None:
                    new_info = zipfile.ZipInfo(item.filename, item.date_time)
                    new_info.compress_type = zipfile.ZIP_DEFLATED
                    zout.writestr(new_info, content)


# ──────────────────────────────────────────────────────────────────────────────
# Header definitions (single source of truth for all output columns)
# ──────────────────────────────────────────────────────────────────────────────

OUTLOOKS_HEADERS = ["email", "email_password", "dob", "prefecture", "nickname", "status", "error_details"]
CATCHALL_HEADERS = ["email", "email_password", "otp_email", "otp_pass", "dob", "prefecture", "nickname", "status", "error_details"]
ACCOUNTS_HEADERS = [
    "email", "bandai_password", "namco_password", "nickname",
    "phone", "bnid_user_code", "proxy_used", "status", "created_at", "error_details"
]
PROXIES_HEADERS  = ["proxy", "status"]

# Map tên cột ngoài lệ (cũ, tiếng Anh viết khác) về key chuẩn của code
COLUMN_ALIASES = {
    "bnid":            "bnid_user_code",
    "bnid user code":  "bnid_user_code",

    "proxy used":      "proxy_used",
    "bandai password": "bandai_password",
    "namco password":  "namco_password",
    "created at":      "created_at",
    "error details":   "error_details",
    "data usage (mb)": "data_usage_mb",
    "dob":             "dob",
    "location":        "prefecture",
    "token":           "ms_token",
    "ms token":        "ms_token",
    "ms_token":        "ms_token",
    "uuid":            "ms_uuid",
    "ms uuid":         "ms_uuid",
    "ms_uuid":         "ms_uuid",
}

# Chỉ các giá trị status hợp lệ hiển thị trong sheet
_STATUS_MAP = {
    "success":    "SUCCESS",
    "failed":     "FAILED",
    "fail_no_retry": "FAIL_NO_RETRY",
    "processing": "PROCESSING",
    "pending":    "PENDING",
    "has_bnid":   "HAS_BNID",
    "aborted":    "FAIL_NO_RETRY",
    "error":      "FAILED",
}

def _normalize_status(status: str) -> str:
    """Chuẩn hóa status về các giá trị hợp lệ."""
    return _STATUS_MAP.get(str(status or "").strip().lower(), "FAILED")


class XlsxConnection:

    def _atomic_save(self, wb, path):
        import os, time
        tmp_path = str(path) + f".tmp_{int(time.time() * 1000)}"
        try:
            wb.save(tmp_path)
            # Kiểm tra kiểm định (Validation): Đảm bảo file .tmp ghi ra hoàn toàn hợp lệ 100% trước khi đổi tên
            test_wb = openpyxl.load_workbook(tmp_path, read_only=True)
            test_wb.close()
            # Thực thi lệnh đổi tên Atomic OS operation (an toàn tuyệt đối)
            os.replace(tmp_path, str(path))
        except Exception as e:
            if os.path.exists(tmp_path):
                try: os.remove(tmp_path)
                except Exception: pass
            try:
                wb.save(str(path))
            except Exception as direct_err:
                log.error(f"❌ LỖI LƯU FILE EXCEL: Không thể lưu file '{path}'. File đang bị mở/khóa bởi Microsoft Excel hoặc ứng dụng khác. Vui lòng TẮT file Excel trước khi chạy! Chi tiết: {direct_err}")
                raise RuntimeError(f"File Excel đang bị khóa bởi phần mềm khác: {direct_err}")

    def _create_session_backup(self):
        """Tự động tạo bản sao lưu an toàn khi khởi động bot."""
        if not self.xlsx_path or not self.xlsx_path.exists():
            return
        try:
            import shutil, os
            backup_dir = self.xlsx_path.parent / "backups"
            backup_dir.mkdir(exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = backup_dir / f"{self.xlsx_path.stem}_backup_{timestamp}{self.xlsx_path.suffix}"
            shutil.copy2(self.xlsx_path, backup_file)
            log.info(f"🛡️ [Cơ chế An toàn] Đã tự động tạo bản sao lưu bảo vệ dữ liệu tại: {backup_file}")
            
            # Xóa các bản backup cũ, chỉ giữ lại 10 file gần nhất
            backups = sorted(backup_dir.glob(f"{self.xlsx_path.stem}_backup_*"), key=os.path.getmtime)
            if len(backups) > 10:
                for b in backups[:-10]:
                    try: os.remove(b)
                    except Exception: pass
        except Exception as e:
            log.warning(f"Không thể tạo bản sao lưu an toàn: {e}")

    def _load_workbook(self, read_only: bool = False):
        """Mở workbook với cơ chế tự động khôi phục nếu file bị Bad CRC-32 hoặc hỏng cấu trúc Zip."""
        if not self.xlsx_path:
            raise RuntimeError("Chưa cấu hình đường dẫn file XLSX.")
        path_str = str(self.xlsx_path)
        try:
            if read_only:
                return openpyxl.load_workbook(path_str, read_only=True)
            return openpyxl.load_workbook(path_str)
        except Exception as e:
            err_msg = str(e)
            if "Bad CRC-32" in err_msg or "BadZipFile" in err_msg or "not a zip file" in err_msg.lower() or "zipfile" in err_msg.lower():
                log.error(f"❌ PHÁT HIỆN FILE EXCEL BỊ HỎNG CẤU TRÚC (Bad CRC-32): {path_str} ({e})")
                log.warning("🔄 Đang tự động sửa và khôi phục dữ liệu từ file Excel bị hỏng...")

                try:
                    import shutil, time
                    backup_name = f"{self.xlsx_path.stem}_CORRUPTED_{int(time.time())}{self.xlsx_path.suffix}"
                    backup_path = self.xlsx_path.parent / backup_name
                    shutil.copy2(self.xlsx_path, backup_path)
                    log.info(f"📦 Đã tạo bản sao lưu file bị hỏng tại: {backup_path}")
                except Exception as b_err:
                    log.warning(f"Không thể sao lưu file hỏng: {b_err}")

                try:
                    repaired_tmp = str(self.xlsx_path) + ".repaired"
                    repair_xlsx_crc(path_str, repaired_tmp)
                    import os
                    os.replace(repaired_tmp, path_str)
                    log.info(f"✅ ĐÃ KHÔI PHỤC THÀNH CÔNG FILE EXCEL BỊ HỎNG CRC: {path_str}")
                    if read_only:
                        return openpyxl.load_workbook(path_str, read_only=True)
                    return openpyxl.load_workbook(path_str)
                except Exception as repair_err:
                    log.error(f"Không thể sửa tự động file Excel: {repair_err}")
            raise e

    def __init__(self, xlsx_path: str):
        self.xlsx_path = Path(xlsx_path) if xlsx_path else None
        self._lock = threading.Lock()
        self._connected = False

        if self.xlsx_path and self.xlsx_path.exists():
            try:
                self._create_session_backup()
                wb = self._load_workbook()
                self._ensure_sheets(wb)
                self._atomic_save(wb, self.xlsx_path)
                wb.close()
                self._connected = True
                log.info(f"✅ XlsxConnection: Kết nối thành công → {self.xlsx_path}")
            except Exception as e:
                log.error(f"❌ XlsxConnection: Không thể mở file XLSX: {e}")


        else:
            if self.xlsx_path:
                log.warning(f"⚠️ XlsxConnection: File không tồn tại: {self.xlsx_path}")
            else:
                log.warning("⚠️ XlsxConnection: Chưa cấu hình đường dẫn file XLSX.")

    # ──────────────────────────────────────────────────────────────────────────
    # Internal helpers
    # ──────────────────────────────────────────────────────────────────────────

    def _get_headers(self, ws):
        """Lấy list headers từ dòng đầu tiên của sheet."""
        return [str(cell.value or "").strip() for cell in ws[1]]

    def _col_index(self, headers, col_name):
        """
        Tìm index của cột dựa trên header name, xử lý cả aliases.
        """
        normalized = [h.lower() for h in headers]
        col_name_norm = col_name.lower()
        
        # Thử tìm trực tiếp
        if col_name_norm in normalized:
            return normalized.index(col_name_norm)
        
        # Thử tìm theo alias
        target = COLUMN_ALIASES.get(col_name_norm)
        if target and target in normalized:
            return normalized.index(target)
            
        raise KeyError(f"Không tìm thấy cột '{col_name}' trong sheet. Headers hiện tại: {headers}")

    def _ensure_sheets(self, wb):
        """Đảm bảo các sheet cơ bản tồn tại."""
        required = ["Outlooks", "Gmails", "Iclouds", "Accounts", "Proxies"]
        for name in required:
            if name not in wb.sheetnames:
                wb.create_sheet(name)
                # Init headers
                ws = wb[name]
                if name == "Accounts":
                    ws.append(ACCOUNTS_HEADERS)
                elif name == "Proxies":
                    ws.append(PROXIES_HEADERS)

    # ──────────────────────────────────────────────────────────────────────────
    # Public: lifecycle
    # ──────────────────────────────────────────────────────────────────────────

    def is_connected(self) -> bool:
        return self._connected

    def reset_interrupted_to_pending(self):
        """
        Gọi khi khởi động bot. Thực hiện 2 việc:
        1. Reset PROCESSING → PENDING  (bị dừng giữa chừng)
        2. Reset FAILED → PENDING      nếu email chưa có BNID trong sheet Accounts
           (FAILED + có BNID → giữ nguyên HAS_BNID để chạy luồng login)
        """
        with self._lock:
            try:
                wb = self._load_workbook()


                # Đọc tập hợp email đã có BNID từ sheet Accounts (check cột bnid_user_code)
                emails_with_bnid: set = set()
                if "Accounts" in wb.sheetnames:
                    ws_acc = wb["Accounts"]
                    acc_headers = self._get_headers(ws_acc)
                    try:
                        email_col   = self._col_index(acc_headers, "email")
                        bnid_col    = self._col_index(acc_headers, "bnid_user_code")
                        for row in ws_acc.iter_rows(min_row=2, values_only=True):
                            email_val = str(row[email_col] or "").strip().split("|")[0].strip()
                            bnid_val  = str(row[bnid_col] or "").strip()
                            if bnid_val:
                                emails_with_bnid.add(email_val.lower())
                    except Exception:
                        pass

                # Xử lý sheet đang active
                active_sheet_name = getattr(config, "ACTIVE_SHEET", "Outlooks")
                if active_sheet_name not in wb.sheetnames:
                    active_sheet_name = "Outlooks" if "Outlooks" in wb.sheetnames else "Mails"
                
                if active_sheet_name in wb.sheetnames:
                    ws_mail = wb[active_sheet_name]
                    mail_headers = self._get_headers(ws_mail)
                    status_col = self._col_index(mail_headers, "status")
                    email_col  = self._col_index(mail_headers, "email")
    
                    reset_proc = 0
                    reset_fail = 0
                    for row in ws_mail.iter_rows(min_row=2):
                        val = str(row[status_col].value or "").strip()
                        norm_val = _normalize_status(val)
                        raw_cell = str(row[email_col].value or "").strip()
                        email_plain = raw_cell.split("|")[0].strip().lower()
    
                        if norm_val == "PROCESSING":
                            row[status_col].value = "PENDING"
                            reset_proc += 1
                        elif norm_val == "FAILED":
                            if email_plain not in emails_with_bnid:
                                # Chưa có BNID → cho phép chạy lại
                                row[status_col].value = "PENDING"
                                reset_fail += 1
                            else:
                                # Đã có BNID → đổi sang HAS_BNID để chạy luồng login
                                row[status_col].value = "HAS_BNID"
                                reset_fail += 1

                if reset_proc + reset_fail > 0:
                    self._atomic_save(wb, self.xlsx_path)
                if reset_proc:
                    log.info(f"♻️  Reset {reset_proc} email PROCESSING → PENDING.")
                if reset_fail:
                    log.info(f"♻️  Reset {reset_fail} email FAILED (chưa có BNID) → PENDING để thử lại.")
                wb.close()
            except Exception as e:
                log.error(f"❌ Lỗi reset interrupted emails: {e}")

    # Alias cũ để không break nếu có code nào khác gọi
    reset_processing_to_pending = reset_interrupted_to_pending

    # ──────────────────────────────────────────────────────────────────────────
    # Public: Input sheets (Outlooks, Gmails, Iclouds)
    # ──────────────────────────────────────────────────────────────────────────

    def get_pending_emails(self, batch_size: int = 50) -> list:
        """
        Đọc tối đa `batch_size` email có status PENDING (hoặc ô trống) từ sheet đang active.

        Cột email hỗ trợ 2 format:
          - Chỉ email:               user@hotmail.com
          - Pipe-separated (1 cột):  user@hotmail.com|password|token|uuid
            Field 1: email
            Field 2: email_password
            Field 3: ms_token (Microsoft refresh token, tùy chọn)
            Field 4: ms_uuid  (Microsoft account UUID, tùy chọn)

        Trả về list[dict].
        """
        with self._lock:
            try:
                wb = self._load_workbook()

                active_sheet_name = getattr(config, "ACTIVE_SHEET", "Outlooks")
                if active_sheet_name not in wb.sheetnames:
                    active_sheet_name = "Outlooks" if "Outlooks" in wb.sheetnames else "Mails"
                    
                if active_sheet_name not in wb.sheetnames:
                    log.error(f"❌ Không tìm thấy sheet {active_sheet_name} trong file XLSX.")
                    return []
                    
                ws = wb[active_sheet_name]
                headers = self._get_headers(ws)
                try:
                    status_col_idx = self._col_index(headers, "status")
                except KeyError:
                    status_col_idx = None

                results = []
                for row in ws.iter_rows(min_row=2):
                    if len(results) >= batch_size:
                        break
                    row_values = [cell.value for cell in row]
                    row_dict = dict(zip(headers, row_values))
                    raw_email_cell = str(row_dict.get("email", "") or "").strip()
                    if not raw_email_cell:
                        continue
                    status = str(row_dict.get("status", "") or "").strip().upper()
                    if status in ("", "PENDING", "HAS_BNID"):
                        # Parse pipe-separated format
                        parts = raw_email_cell.split("|")
                        email         = parts[0].strip()
                        email_password = parts[1].strip() if len(parts) > 1 else str(row_dict.get("email_password", "") or "").strip()
                        ms_token      = parts[2].strip() if len(parts) > 2 else ""
                        ms_uuid       = parts[3].strip() if len(parts) > 3 else ""

                        results.append({
                            "email": email,
                            "raw_email": raw_email_cell,  # giữ nguyên giá trị gốc để update status
                            "email_password": email_password,
                            "ms_token": ms_token,
                            "ms_uuid": ms_uuid,
                            "otp_email": str(row_dict.get("otp_email", "") or "").strip(),
                            "otp_pass": str(row_dict.get("otp_pass", "") or "").strip(),
                            "dob": str(row_dict.get("dob", "") or "").strip(),
                            "prefecture": str(row_dict.get("prefecture", "") or "").strip(),
                            "nickname": str(row_dict.get("nickname", "") or "").strip(),
                            "provider": active_sheet_name.lower(),
                            "mail_status": status,  # status gốc từ Mails sheet (PENDING/HAS_BNID)
                        })
                self._atomic_save(wb, self.xlsx_path)
                wb.close()
                log.info(f"📋 Đọc được {len(results)} email PENDING từ XLSX.")
                return results
            except Exception as e:
                log.error(f"❌ Lỗi đọc email từ XLSX: {e}")
                return []


    def update_email_status(self, email: str, status: str, error_details: str = "", extra_data: dict = None):
        """
        Cập nhật cột status và error_details trong sheet active.
        Tìm dòng theo email (hoặc raw pipe string nếu truyền vào).
        """
        email = str(email or "").strip()
        if not email:
            return
        with self._lock:
            try:
                wb = self._load_workbook()

                active_sheet_name = getattr(config, "ACTIVE_SHEET", "Outlooks")
                if active_sheet_name not in wb.sheetnames:
                    active_sheet_name = "Outlooks" if "Outlooks" in wb.sheetnames else "Mails"
                
                if active_sheet_name not in wb.sheetnames:
                    return
                    
                ws = wb[active_sheet_name]
                headers = self._get_headers(ws)
                status_col = self._col_index(headers, "status")
                email_col  = self._col_index(headers, "email")

                try:
                    err_col = self._col_index(headers, "error_details")
                except KeyError:
                    err_col = len(headers)
                    ws.cell(1, err_col + 1, "error_details")
                    headers.append("error_details")

                for row in ws.iter_rows(min_row=2):
                    cell_val = str(row[email_col].value or "").strip()
                    # Match theo raw value (pipe string) hoặc chỉ phần email trước |
                    cell_email = cell_val.split("|")[0].strip()
                    if cell_val == email or cell_email == email:
                        row[status_col].value = _normalize_status(status)
                        if error_details:
                            row[err_col].value = error_details
                        break


                self._atomic_save(wb, self.xlsx_path)
                wb.close()
                log.debug(f"📝 Cập nhật status '{status}' cho: {email}")
            except Exception as e:
                log.error(f"❌ Lỗi cập nhật email status: {e}")


    # ──────────────────────────────────────────────────────────────────────────
    # Public: Accounts sheet
    # ──────────────────────────────────────────────────────────────────────────

    def append_account(self, data: dict):
        """
        Upsert kết quả vào sheet Accounts theo email.
        Dung ws.cell(row, column) de dam bao dung cot, khong bi lech.
        """
        email = str(data.get("email", "") or "").strip()
        if not email:
            return

        status = _normalize_status(data.get("status", ""))
        # CHỈ CHO PHÉP ghi vào Accounts sheet nếu SUCCESS hoặc HAS_BNID
        if status not in ("SUCCESS", "HAS_BNID"):
            return

        if not data.get("created_at") and data.get("status") in ("SUCCESS",):
            data["created_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Chuan hoa status chi cac gia tri hop le: SUCCESS / FAILED / PROCESSING / PENDING / HAS_BNID
        data["status"] = _normalize_status(data.get("status", ""))

        with self._lock:
            try:
                wb = self._load_workbook()

                
                # Tìm sheet Accounts (case-insensitive hoặc tự tạo)
                target_sheet_name = None
                for sname in wb.sheetnames:
                    if sname.strip().lower() in ("accounts", "account"):
                        target_sheet_name = sname
                        break
                
                if not target_sheet_name:
                    ws = wb.create_sheet("Accounts")
                    ws.append(ACCOUNTS_HEADERS)
                    target_sheet_name = "Accounts"
                else:
                    ws = wb[target_sheet_name]

                sheet_headers = self._get_headers(ws)
                if not sheet_headers or "email" not in sheet_headers:
                    ws.cell(row=1, column=1, value="email")
                    sheet_headers = self._get_headers(ws)

                # Thêm cột còn thiếu vào cuối sheet
                for col_name in ACCOUNTS_HEADERS:
                    if col_name not in sheet_headers:
                        new_col_num = len(sheet_headers) + 1  # 1-indexed
                        ws.cell(row=1, column=new_col_num, value=col_name)
                        sheet_headers.append(col_name)

                # col_map: ten cot -> so cot 1-indexed
                col_map = {h: idx + 1 for idx, h in enumerate(sheet_headers)}
                email_col_num = col_map.get("email", 1)

                # Tìm dòng có email trùng
                target_row_num = None
                for row in ws.iter_rows(min_row=2):
                    cell_val = str(ws.cell(row=row[0].row, column=email_col_num).value or "").strip()
                    if cell_val == email:
                        target_row_num = row[0].row
                        break

                if target_row_num is not None:
                    # UPDATE: dùng ws.cell(row, col) để ghi đúng vị trí
                    for key, val in data.items():
                        if key in col_map:
                            col_num = col_map[key]
                            existing = ws.cell(row=target_row_num, column=col_num).value
                            if val != "" or key in ("status", "error_details"):
                                ws.cell(row=target_row_num, column=col_num, value=val)
                            elif existing is None:
                                ws.cell(row=target_row_num, column=col_num, value="")
                    log.debug(f"Update: {email} -> status={data['status']} | phone={data.get('phone','')}")
                else:
                    # INSERT: dùng ws.cell(row, col) để ghi mới
                    next_row = ws.max_row + 1
                    for key, val in data.items():
                        if key in col_map:
                            ws.cell(row=next_row, column=col_map[key], value=val or "")
                    log.debug(f"Insert: {email} -> status={data['status']} | phone={data.get('phone','')}")


                self._atomic_save(wb, self.xlsx_path)
                wb.close()
            except Exception as e:
                log.error(f"❌ [LỖI GHI SHEET ACCOUNTS] Không thể ghi email {email} vào file Excel: {e}", exc_info=True)



    def get_account_status(self, email: str) -> str:
        """Kiểm tra email đã có trong Accounts chưa, trả về status hoặc rỗng."""
        info = self.get_account_info(email)
        return info.get("status", "")

    def get_account_info(self, email: str) -> dict:
        """Lấy thông tin account từ sheet Accounts: status, bandai_password."""
        email = str(email or "").strip()
        with self._lock:
            try:
                wb = self._load_workbook(read_only=True)

                ws = wb["Accounts"]
                headers = self._get_headers(ws)
                email_col  = self._col_index(headers, "email")
                status_col = self._col_index(headers, "status")
                pass_col   = self._col_index(headers, "bandai_password") if "bandai_password" in headers else None
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if str(row[email_col] or "").strip() == email:
                        result = {
                            "status": str(row[status_col] or "").strip(),
                            "bandai_password": str(row[pass_col] or "").strip() if pass_col is not None else ""
                        }
                        wb.close()
                        return result
                wb.close()
                return {"status": "", "bandai_password": ""}
            except Exception as e:
                log.error(f"❌ Lỗi đọc account info: {e}")
                return {"status": "", "bandai_password": ""}

    # ──────────────────────────────────────────────────────────────────────────
    # Public: Proxies sheet
    # ──────────────────────────────────────────────────────────────────────────

    def get_active_proxies(self) -> list:
        """Đọc danh sách proxy active từ sheet Proxies. Trả về list[str]."""
        with self._lock:
            try:
                wb = self._load_workbook(read_only=True)

                if "Proxies" not in wb.sheetnames:
                    wb.close()
                    return []
                ws = wb["Proxies"]
                headers = self._get_headers(ws)
                proxy_col  = self._col_index(headers, "proxy")
                status_col = self._col_index(headers, "status") if "status" in headers else None

                results = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    proxy = str(row[proxy_col] or "").strip()
                    if not proxy:
                        continue
                    if status_col is not None:
                        status = str(row[status_col] or "").strip().lower()
                        if status in ("disabled", "inactive", "0"):
                            continue
                    results.append(proxy)
                wb.close()
                log.info(f"🔌 Đọc được {len(results)} proxy active từ XLSX.")
                return results
            except Exception as e:
                log.error(f"❌ Lỗi đọc proxies: {e}")
                return []

    def load_permanent_counts(self, proxy_pool):
        """
        Đếm số SUCCESS theo proxy từ sheet Accounts để khôi phục trạng thái proxy_pool.
        """
        with self._lock:
            try:
                wb = self._load_workbook(read_only=True)

                ws = wb["Accounts"]
                headers = self._get_headers(ws)
                proxy_col  = self._col_index(headers, "proxy_used")
                status_col = self._col_index(headers, "status")

                counts: dict = {}
                for row in ws.iter_rows(min_row=2, values_only=True):
                    status = str(row[status_col] or "").strip()
                    if status == "SUCCESS":
                        proxy = str(row[proxy_col] or "").strip()
                        if proxy:
                            counts[proxy] = counts.get(proxy, 0) + 1
                wb.close()

                for proxy_str, count in counts.items():
                    if hasattr(proxy_pool, "set_permanent_count"):
                        proxy_pool.set_permanent_count(proxy_str, count)
                log.info(f"📊 Khôi phục permanent count cho {len(counts)} proxy từ XLSX.")
            except Exception as e:
                log.warning(f"⚠️ Không thể load permanent counts: {e}")

    # ──────────────────────────────────────────────────────────────────────────
    # Static: tạo file mẫu
    # ──────────────────────────────────────────────────────────────────────────

    @staticmethod
    def create_template(xlsx_path: str) -> bool:
        """Tạo file XLSX mẫu với các sheet: Outlooks, Gmails, Iclouds, Accounts, Proxies."""
        try:
            wb = Workbook()

            # Sheet Outlooks
            ws_outlooks = wb.active
            ws_outlooks.title = "Outlooks"
            XlsxConnection._write_header(ws_outlooks, OUTLOOKS_HEADERS, color="4472C4")
            ws_outlooks.append(["example@hotmail.com", "emailpassword", "1995-06-15", "東京都", "", "PENDING"])
            for col, w in zip("ABCDEF", [35, 20, 14, 16, 20, 14]):
                ws_outlooks.column_dimensions[col].width = w

            # Sheet Gmails
            ws_gmails = wb.create_sheet("Gmails")
            XlsxConnection._write_header(ws_gmails, CATCHALL_HEADERS, color="D32F2F")
            ws_gmails.append(["alias@gmail.com", "mainpass", "catchall@gmail.com", "app_password_here", "1995-06-15", "東京都", "", "PENDING"])
            for i, w in enumerate([35, 20, 35, 25, 14, 16, 20, 14], 1):
                ws_gmails.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

            # Sheet Iclouds
            ws_iclouds = wb.create_sheet("Iclouds")
            XlsxConnection._write_header(ws_iclouds, CATCHALL_HEADERS, color="0288D1")
            ws_iclouds.append(["alias@icloud.com", "mainpass", "catchall@icloud.com", "app_password_here", "1995-06-15", "東京都", "", "PENDING"])
            for i, w in enumerate([35, 20, 35, 25, 14, 16, 20, 14], 1):
                ws_iclouds.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

            # Sheet Accounts
            ws_accounts = wb.create_sheet("Accounts")
            XlsxConnection._write_header(ws_accounts, ACCOUNTS_HEADERS, color="70AD47")
            for i, w in enumerate([35, 18, 18, 20, 16, 18, 30, 12, 20, 40], 1):
                ws_accounts.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

            # Sheet Proxies
            ws_proxies = wb.create_sheet("Proxies")
            XlsxConnection._write_header(ws_proxies, PROXIES_HEADERS, color="ED7D31")
            ws_proxies.append(["http://host:port:user:pass", "active"])
            ws_proxies.column_dimensions["A"].width = 40
            ws_proxies.column_dimensions["B"].width = 12

            import os; tmp = str(xlsx_path)+".tmp"; wb.save(tmp); os.replace(tmp, str(xlsx_path))
            wb.close()
            log.info(f"✅ Đã tạo file XLSX mẫu tại: {xlsx_path}")
            return True
        except Exception as e:
            log.error(f"❌ Lỗi tạo file XLSX mẫu: {e}")
            return False

    # ──────────────────────────────────────────────────────────────────────────
    # Private helpers
    # ──────────────────────────────────────────────────────────────────────────

    def _ensure_sheets(self, wb: Workbook):
        """Tạo các sheet còn thiếu trong file XLSX."""
        if "Outlooks" not in wb.sheetnames and "Mails" not in wb.sheetnames:
            ws = wb.create_sheet("Outlooks")
            self._write_header(ws, OUTLOOKS_HEADERS, color="4472C4")
        if "Gmails" not in wb.sheetnames:
            ws = wb.create_sheet("Gmails")
            self._write_header(ws, CATCHALL_HEADERS, color="D32F2F")
        if "Iclouds" not in wb.sheetnames:
            ws = wb.create_sheet("Iclouds")
            self._write_header(ws, CATCHALL_HEADERS, color="0288D1")
        if "Accounts" not in wb.sheetnames:
            ws = wb.create_sheet("Accounts")
            self._write_header(ws, ACCOUNTS_HEADERS, color="70AD47")
        if "Proxies" not in wb.sheetnames:
            ws = wb.create_sheet("Proxies")
            self._write_header(ws, PROXIES_HEADERS, color="ED7D31")

    @staticmethod
    def _write_header(ws, headers: list, color: str = "4472C4"):
        ws.append(headers)
        for cell in ws[1]:
            cell.font  = Font(bold=True, color="FFFFFF")
            cell.fill  = PatternFill("solid", fgColor=color)
            cell.alignment = Alignment(horizontal="center")

    @staticmethod
    def _get_headers(ws) -> list:
        """Trả về headers của sheet, đã áp alias về key chuẩn."""
        raw = [str(cell.value or "").strip().lower() for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        return [COLUMN_ALIASES.get(h, h) for h in raw]

    @staticmethod
    def _col_index(headers: list, col_name: str) -> int:
        col_name = col_name.lower()
        try:
            return headers.index(col_name)
        except ValueError:
            raise KeyError(f"Không tìm thấy cột '{col_name}'. Headers: {headers}")
